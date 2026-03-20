import openpyxl
import sys

wb = openpyxl.load_workbook(r'F:\Chess_Planner\Thời Khóa Biểu Chi Tiết Hàng Tuần.xlsx')

with open(r'F:\Chess_Planner\excel_output.txt', 'w', encoding='utf-8') as f:
    f.write(f'Sheet names: {wb.sheetnames}\n')
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        f.write(f'\n=== SHEET: {sheet_name} === (rows={ws.max_row}, cols={ws.max_column})\n')
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
            vals = []
            for cell in row:
                v = cell.value
                if v is not None:
                    vals.append(str(v))
                else:
                    vals.append('')
                    
            f.write(' | '.join(vals) + '\n')

print("Done")
